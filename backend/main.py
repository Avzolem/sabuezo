"""Sabuezo backend — FastAPI server.

Endpoints:
  POST /analyze/text   → analiza mensaje de texto
  POST /analyze/image  → analiza screenshot
  POST /analyze/url    → analiza un link
  POST /scan           → escaneo de seguridad del sitio
  POST /pyme/register  → registra PyME
  GET  /health
"""
import os
from dotenv import load_dotenv
from pathlib import Path

# Carga .env desde la raíz del proyecto (un nivel arriba de backend/)
load_dotenv(Path(__file__).parent.parent / ".env")

import io
from fastapi import FastAPI, Header, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional

from analyzers import text as text_analyzer
from analyzers import image as image_analyzer
from analyzers import url as url_analyzer
from analyzers import scanner as site_scanner
from analyzers import breaches as breach_checker
import db

INTERNAL_TOKEN = os.getenv("INTERNAL_API_TOKEN", "dev-token")

app = FastAPI(title="Sabuezo API", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def require_internal(x_internal_token: Optional[str] = Header(None)):
    if x_internal_token != INTERNAL_TOKEN:
        raise HTTPException(status_code=401, detail="Invalid internal token")


class TextRequest(BaseModel):
    user_id: str
    text: str
    pushname: Optional[str] = None


class ImageRequest(BaseModel):
    user_id: str
    image_base64: str
    caption: Optional[str] = ""
    pushname: Optional[str] = None


class UrlRequest(BaseModel):
    user_id: str
    url: str
    pushname: Optional[str] = None


class ScanRequest(BaseModel):
    url: str
    owner_email: Optional[str] = None
    user_id: Optional[str] = None
    pushname: Optional[str] = None


class BreachCheckRequest(BaseModel):
    user_id: Optional[str] = None
    value: str
    source: Optional[str] = None  # 'web' | 'bot' (si falta, se infiere)


class PymeRegisterRequest(BaseModel):
    owner_jid: str
    name: str
    website: str
    owner_email: Optional[str] = None
    pushname: Optional[str] = None


class LidMapRequest(BaseModel):
    lid: str
    phone_jid: Optional[str] = None
    pushname: Optional[str] = None


def _persist_detection_safe(user_jid, kind, result, raw_content="", pushname=None):
    """Wrap en try para que un fallo de DB no rompa la respuesta al usuario.

    Además: si la PyME tiene scan reciente con fallas SPF/DMARC y el mensaje
    es phishing (rojo/amarillo), enriquece `result` con `cross_insight`.
    """
    try:
        pyme = db.get_pyme_by_jid(user_jid)
        pyme_id = pyme["id"] if pyme else None
        db.save_phishing_detection(
            user_jid=user_jid,
            kind=kind,
            risk=result.get("risk", "amarillo"),
            analysis=result,
            raw_content=raw_content,
            pushname=pushname,
            pyme_id=pyme_id,
        )

        # Cross-cutting insight: conecta el phishing entrante con la postura
        # de seguridad del propio dominio (SPF/DMARC). Solo si hay sospecha.
        if pyme_id and result.get("risk") in ("rojo", "amarillo"):
            scan = db.latest_scan_for_pyme(pyme_id)
            if scan:
                email_auth = (scan.get("raw") or {}).get("email_auth") or {}
                spf_ok = bool(email_auth.get("spf_present"))
                dmarc_ok = bool(email_auth.get("dmarc_present"))
                dmarc_policy = (email_auth.get("dmarc_policy") or "").lower()
                # DMARC con p=none es casi como no tenerlo
                dmarc_weak = dmarc_ok and dmarc_policy == "none"
                if (not spf_ok) or (not dmarc_ok) or dmarc_weak:
                    domain = scan.get("domain") or pyme.get("website") or "tu dominio"
                    missing = []
                    if not spf_ok:
                        missing.append("SPF ausente")
                    if not dmarc_ok:
                        missing.append("DMARC ausente")
                    elif dmarc_weak:
                        missing.append("DMARC en modo p=none (no bloquea)")
                    result["cross_insight"] = {
                        "domain": domain,
                        "missing": missing,
                        "scan_id": scan.get("id"),
                        "message": (
                            f"Tu dominio *{domain}* tiene {', '.join(missing)}. "
                            "Eso significa que cualquiera puede mandar correos "
                            "haciéndose pasar por tu empresa — justamente como "
                            "este mensaje. Revisa tu reporte de seguridad para "
                            "ver cómo cerrar esa puerta."
                        ),
                    }
    except Exception as e:
        print(f"[db] error guardando detección: {e}")


def _persist_breach_safe(kind, value, result, source=None, user_jid=None):
    """Registra el chequeo de filtración sin romper la respuesta si la DB falla.

    Solo persiste consultas válidas (result.ok). El `domain` es el dominio del
    correo o el código de país del teléfono — útil para métricas agregadas.
    """
    try:
        if not result.get("ok"):
            return
        if kind == "email":
            domain = value.split("@")[-1].lower() if "@" in value else None
        else:
            domain = value[:2] if value else None  # código de país (52, 57, …)
        db.save_breach_check(
            kind=kind,
            value=value,
            found=bool(result.get("found")),
            breach_count=int(result.get("count") or 0),
            domain=domain,
            source=source,
            user_jid=user_jid,
        )
    except Exception as e:
        print(f"[db] error guardando breach_check: {e}")


@app.get("/health")
async def health():
    return {"status": "ok", "service": "sabuezo"}


@app.get("/metrics")
async def metrics(x_internal_token: Optional[str] = Header(None)):
    require_internal(x_internal_token)
    return db.compute_metrics()


@app.get("/metrics/detail")
async def metrics_detail(x_internal_token: Optional[str] = Header(None)):
    """Listas completas (valores en claro) de correos y teléfonos consultados."""
    require_internal(x_internal_token)
    rows = db.breach_rows()
    return {
        "correos": [r for r in rows if r.get("kind") == "email"],
        "telefonos": [r for r in rows if r.get("kind") == "phone"],
    }


@app.get("/metrics/export.xlsx")
async def metrics_export(x_internal_token: Optional[str] = Header(None)):
    """Genera un Excel con 3 hojas: Correos, Teléfonos y Sitios escaneados."""
    require_internal(x_internal_token)
    from openpyxl import Workbook
    from openpyxl.styles import Font

    rows = db.breach_rows()
    scans = db.scan_rows()

    def header(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Correos"
    header(ws1, ["Correo", "Dominio", "Filtrado", "Nº filtraciones", "Fuente", "Fecha"])
    for r in (x for x in rows if x.get("kind") == "email"):
        ws1.append([
            r.get("value"), r.get("domain"),
            "Sí" if r.get("found") else "No",
            r.get("breach_count") or 0, r.get("source"),
            (r.get("created_at") or "")[:19].replace("T", " "),
        ])

    ws2 = wb.create_sheet("Teléfonos")
    header(ws2, ["Teléfono", "País", "Filtrado", "Nº filtraciones", "Fuente", "Fecha"])
    for r in (x for x in rows if x.get("kind") == "phone"):
        ws2.append([
            r.get("value"), r.get("domain"),
            "Sí" if r.get("found") else "No",
            r.get("breach_count") or 0, r.get("source"),
            (r.get("created_at") or "")[:19].replace("T", " "),
        ])

    ws3 = wb.create_sheet("Sitios escaneados")
    header(ws3, ["URL", "Dominio", "Score", "Fecha"])
    for r in scans:
        ws3.append([
            r.get("url"), r.get("domain"), r.get("score"),
            (r.get("created_at") or "")[:19].replace("T", " "),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sabuezo-export.xlsx"},
    )


@app.post("/analyze/text")
async def analyze_text(req: TextRequest, x_internal_token: Optional[str] = Header(None)):
    require_internal(x_internal_token)
    result = await text_analyzer.analyze(req.text)
    _persist_detection_safe(req.user_id, "text", result, raw_content=req.text, pushname=req.pushname)
    return result


@app.post("/analyze/image")
async def analyze_image(req: ImageRequest, x_internal_token: Optional[str] = Header(None)):
    require_internal(x_internal_token)
    result = await image_analyzer.analyze(req.image_base64, caption=req.caption or "")
    _persist_detection_safe(req.user_id, "image", result, raw_content=req.caption or "", pushname=req.pushname)
    return result


@app.post("/analyze/url")
async def analyze_url(req: UrlRequest, x_internal_token: Optional[str] = Header(None)):
    require_internal(x_internal_token)
    result = await url_analyzer.analyze(req.url)
    _persist_detection_safe(req.user_id, "url", result, raw_content=req.url, pushname=req.pushname)
    return result


@app.post("/pyme/register")
async def pyme_register(req: PymeRegisterRequest, x_internal_token: Optional[str] = Header(None)):
    require_internal(x_internal_token)
    try:
        pyme = db.upsert_pyme(
            owner_jid=req.owner_jid,
            name=req.name,
            website=req.website,
            owner_email=req.owner_email,
            pushname=req.pushname,
        )
        return {"ok": True, "pyme": pyme}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/lid-map")
async def lid_map(req: LidMapRequest, x_internal_token: Optional[str] = Header(None)):
    require_internal(x_internal_token)
    try:
        row = db.upsert_lid_map(
            lid=req.lid,
            phone_jid=req.phone_jid,
            pushname=req.pushname,
        )
        return {"ok": True, "lid_map": row}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/check/email")
async def check_email_breach(req: BreachCheckRequest, x_internal_token: Optional[str] = Header(None)):
    require_internal(x_internal_token)
    value = (req.value or "").strip().lower()
    result = await breach_checker.check_email(value)
    source = req.source or ("bot" if req.user_id else "web")
    _persist_breach_safe("email", value, result, source=source, user_jid=req.user_id)
    return result


@app.post("/check/phone")
async def check_phone_breach(req: BreachCheckRequest, x_internal_token: Optional[str] = Header(None)):
    require_internal(x_internal_token)
    value = breach_checker.normalize_phone(req.value)
    result = await breach_checker.check_phone(req.value)
    source = req.source or ("bot" if req.user_id else "web")
    _persist_breach_safe("phone", value, result, source=source, user_jid=req.user_id)
    return result


@app.get("/pyme/by-jid/{jid}/last-scan")
async def pyme_last_scan(jid: str, x_internal_token: Optional[str] = Header(None)):
    require_internal(x_internal_token)
    pyme = db.get_pyme_by_jid(jid)
    if not pyme:
        return {"ok": False, "error": "pyme_not_found"}
    scan = db.latest_scan_for_pyme(pyme["id"])
    return {"ok": True, "pyme": pyme, "scan": scan}


@app.post("/scan")
async def scan_site(req: ScanRequest, x_internal_token: Optional[str] = Header(None)):
    require_internal(x_internal_token)
    result = await site_scanner.scan(req.url, owner_email=req.owner_email)

    # Asociar al PyME si el user ya está registrado
    pyme_id = None
    if req.user_id:
        try:
            pyme = db.get_pyme_by_jid(req.user_id)
            if pyme:
                pyme_id = pyme["id"]
        except Exception as e:
            print(f"[db] error buscando PyME: {e}")

    try:
        saved = db.save_scan(pyme_id, result)
        result["scan_id"] = saved.get("id")
    except Exception as e:
        print(f"[db] error guardando scan: {e}")

    return result


if __name__ == "__main__":
    import uvicorn

    port = int(os.getenv("BACKEND_PORT", "8787"))
    host = os.getenv("BACKEND_HOST", "0.0.0.0")
    uvicorn.run("main:app", host=host, port=port, reload=True)
